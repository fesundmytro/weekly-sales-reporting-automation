#!/usr/bin/env python3
"""Generate a weekly Excel sales report from the newest CSV in input/."""

from __future__ import annotations

import argparse
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = BASE_DIR / "input"
DEFAULT_OUTPUT_DIR = BASE_DIR / "reports"
DEFAULT_LOG_DIR = BASE_DIR / "logs"

REQUIRED_COLUMNS = {
    "order_id",
    "order_date",
    "sales",
    "quantity",
    "profit",
}

CATEGORICAL_COLUMNS = [
    "category",
    "market",
    "region",
    "product_name",
]

NUMERIC_COLUMNS = [
    "sales",
    "quantity",
    "profit",
    "discount",
    "shipping_cost",
]

DATE_COLUMNS = ["order_date", "ship_date"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
KPI_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Clean the newest weekly sales CSV and generate an Excel report."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help="Folder containing weekly CSV files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Folder for generated reports.",
    )
    return parser.parse_args()


def configure_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / "weekly_report.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def normalize_column_name(column: Any) -> str:
    text = str(column)
    text = text.replace("\ufeff", "").replace("ï»¿", "")
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def read_csv_flexible(path: Path) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "latin1"):
        try:
            frame = pd.read_csv(path, encoding=encoding)
            logging.info("Read %s with encoding %s", path.name, encoding)
            return frame
        except UnicodeDecodeError as error:
            last_error = error
    raise ValueError(f"Could not decode {path.name}: {last_error}")


def parse_file_date(path: Path) -> datetime:
    match = re.search(r"(\d{4}-\d{2}-\d{2})", path.stem)
    if match:
        return datetime.strptime(match.group(1), "%Y-%m-%d")
    return datetime.fromtimestamp(path.stat().st_mtime)


def find_weekly_files(input_dir: Path) -> list[Path]:
    input_dir.mkdir(parents=True, exist_ok=True)
    files = sorted(input_dir.glob("*.csv"), key=parse_file_date)
    if not files:
        raise FileNotFoundError(
            f"No CSV files found in {input_dir}. "
            "Add a file named like sales_2026-07-27.csv."
        )
    return files


def add_reject_reason(reasons: pd.Series, mask: pd.Series, label: str) -> None:
    current = reasons.loc[mask]
    reasons.loc[mask] = np.where(current.eq(""), label, current + "; " + label)


def clean_sales_data(raw: pd.DataFrame, source_name: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = raw.copy()
    df.columns = [normalize_column_name(column) for column in df.columns]

    missing_required = sorted(REQUIRED_COLUMNS - set(df.columns))
    if missing_required:
        raise ValueError(
            "Missing required columns: " + ", ".join(missing_required)
        )

    rows_loaded = len(df)
    missing_before = df.isna().sum().sort_values(ascending=False)

    exact_duplicates = int(df.duplicated().sum())
    df = df.drop_duplicates().copy()

    for column in NUMERIC_COLUMNS:
        if column in df.columns:
            cleaned_text = (
                df[column]
                .astype(str)
                .str.replace(",", "", regex=False)
                .str.strip()
                .replace({"": np.nan, "nan": np.nan, "None": np.nan})
            )
            df[column] = pd.to_numeric(cleaned_text, errors="coerce")

    for column in DATE_COLUMNS:
        if column in df.columns:
            df[column] = pd.to_datetime(
                df[column],
                dayfirst=True,
                format="mixed",
                errors="coerce",
            )

    for column in CATEGORICAL_COLUMNS:
        if column not in df.columns:
            df[column] = "Unknown"
        df[column] = (
            df[column]
            .astype("string")
            .str.strip()
            .replace({"": pd.NA})
            .fillna("Unknown")
        )

    reasons = pd.Series("", index=df.index, dtype="string")
    add_reject_reason(reasons, df["order_id"].isna(), "missing order_id")
    add_reject_reason(reasons, df["order_date"].isna(), "invalid order_date")
    add_reject_reason(reasons, df["sales"].isna(), "invalid sales")
    add_reject_reason(reasons, df["quantity"].isna(), "invalid quantity")
    add_reject_reason(reasons, df["profit"].isna(), "invalid profit")
    add_reject_reason(reasons, df["sales"].lt(0), "negative sales")
    add_reject_reason(reasons, df["quantity"].le(0), "quantity must be positive")

    if "discount" in df.columns:
        add_reject_reason(
            reasons,
            df["discount"].notna() & ~df["discount"].between(0, 1),
            "discount outside 0-1",
        )

    rejected = df.loc[reasons.ne("")].copy()
    rejected["reject_reason"] = reasons.loc[reasons.ne("")]

    cleaned = df.loc[reasons.eq("")].copy()
    cleaned = cleaned.sort_values("order_date").reset_index(drop=True)
    rejected = rejected.reset_index(drop=True)

    quality_rows: list[dict[str, Any]] = [
        {"metric": "Source file", "value": source_name},
        {"metric": "Rows loaded", "value": rows_loaded},
        {"metric": "Exact duplicates removed", "value": exact_duplicates},
        {"metric": "Rows rejected", "value": len(rejected)},
        {"metric": "Rows included in report", "value": len(cleaned)},
    ]

    for column, count in missing_before.items():
        if count > 0:
            quality_rows.append(
                {"metric": f"Missing before cleaning: {column}", "value": int(count)}
            )

    quality = pd.DataFrame(quality_rows)
    return cleaned, rejected, quality


def calculate_metrics(df: pd.DataFrame) -> dict[str, float]:
    if df.empty:
        raise ValueError("No valid rows remain after cleaning.")

    total_sales = float(df["sales"].sum())
    total_profit = float(df["profit"].sum())
    order_sales = df.groupby("order_id", dropna=False)["sales"].sum()

    return {
        "Total Sales": total_sales,
        "Total Profit": total_profit,
        "Quantity Sold": float(df["quantity"].sum()),
        "Unique Orders": float(df["order_id"].nunique()),
        "Average Order Value": float(order_sales.mean()),
        "Profit Margin": total_profit / total_sales if total_sales else np.nan,
    }


def grouped_summary(df: pd.DataFrame, group_column: str) -> pd.DataFrame:
    result = (
        df.groupby(group_column, as_index=False, dropna=False)
        .agg(
            total_sales=("sales", "sum"),
            total_profit=("profit", "sum"),
            quantity=("quantity", "sum"),
            unique_orders=("order_id", "nunique"),
        )
        .sort_values("total_sales", ascending=False)
    )
    result["profit_margin"] = np.where(
        result["total_sales"].ne(0),
        result["total_profit"] / result["total_sales"],
        np.nan,
    )
    return result.reset_index(drop=True)


def top_products(df: pd.DataFrame, limit: int = 10) -> pd.DataFrame:
    return (
        df.groupby("product_name", as_index=False, dropna=False)
        .agg(
            total_sales=("sales", "sum"),
            total_profit=("profit", "sum"),
            quantity=("quantity", "sum"),
            unique_orders=("order_id", "nunique"),
        )
        .sort_values("total_sales", ascending=False)
        .head(limit)
        .reset_index(drop=True)
    )


def percent_change(current: float, previous: float | None) -> float | None:
    if previous is None or pd.isna(previous) or previous == 0:
        return None
    return (current - previous) / abs(previous)


def style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def format_worksheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")

    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells[:1000]
        )
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 35)


def write_dataframe_sheet(wb: Workbook, title: str, df: pd.DataFrame, table_name: str) -> None:
    ws = wb.create_sheet(title)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    style_header(ws[1])
    format_worksheet(ws)

    if ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName=table_name, ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center")

    for column_index, column_name in enumerate(df.columns, start=1):
        letter = get_column_letter(column_index)
        if "date" in column_name or column_name in {"order_date", "ship_date", "week"}:
            for cell in ws[letter][1:]:
                cell.number_format = "yyyy-mm-dd"
        elif column_name in {"sales", "profit", "shipping_cost", "total_sales", "total_profit"}:
            for cell in ws[letter][1:]:
                cell.number_format = '#,##0.00'
        elif column_name in {"profit_margin", "discount"}:
            for cell in ws[letter][1:]:
                cell.number_format = "0.00%"
        elif column_name in {"quantity", "unique_orders"}:
            for cell in ws[letter][1:]:
                cell.number_format = '#,##0'


def create_dashboard(
    wb: Workbook,
    current_file: Path,
    current_df: pd.DataFrame,
    current_metrics: dict[str, float],
    previous_file: Path | None,
    previous_metrics: dict[str, float] | None,
) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "Weekly Sales Report"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    period_start = current_df["order_date"].min()
    period_end = current_df["order_date"].max()
    ws["A2"] = "Current file"
    ws["B2"] = current_file.name
    ws["D2"] = "Report period"
    ws["E2"] = f"{period_start:%Y-%m-%d} — {period_end:%Y-%m-%d}"

    headers = ["Metric", "Current", "Previous", "Change"]
    for column, value in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=column, value=value)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    metric_rows: dict[str, int] = {}
    for row_index, (metric, current_value) in enumerate(current_metrics.items(), start=5):
        metric_rows[metric] = row_index
        previous_value = previous_metrics.get(metric) if previous_metrics else None
        change = percent_change(current_value, previous_value)

        ws.cell(row=row_index, column=1, value=metric)
        ws.cell(row=row_index, column=2, value=current_value)
        ws.cell(row=row_index, column=3, value=previous_value)
        ws.cell(row=row_index, column=4, value=change)

        ws.cell(row=row_index, column=1).font = BOLD_FONT
        ws.cell(row=row_index, column=2).fill = KPI_FILL

        for column in range(1, 5):
            ws.cell(row=row_index, column=column).border = THIN_BORDER

    for metric in ("Total Sales", "Total Profit", "Average Order Value"):
        row = metric_rows[metric]
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3).number_format = '#,##0.00'

    for metric in ("Quantity Sold", "Unique Orders"):
        row = metric_rows[metric]
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).number_format = '#,##0'

    margin_row = metric_rows["Profit Margin"]
    ws.cell(row=margin_row, column=2).number_format = "0.00%"
    ws.cell(row=margin_row, column=3).number_format = "0.00%"

    for row in range(5, 11):
        ws.cell(row=row, column=4).number_format = "0.00%"

    ws.conditional_formatting.add(
        "D5:D10",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        "D5:D10",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )

    ws["A12"] = "Previous file"
    ws["B12"] = previous_file.name if previous_file else "Not available"
    ws["A14"] = "Automation impact"
    ws["B14"] = "Manual process: ~3 hours/week"
    ws["B15"] = "Automated run: ~2 minutes"
    ws["B16"] = "Estimated time saved: ~154 hours/year"

    for cell in (ws["A12"], ws["A14"]):
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL

    for column, width in {"A": 24, "B": 22, "C": 18, "D": 16, "E": 24, "F": 12}.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A4"


def add_dashboard_charts(wb: Workbook) -> None:
    dashboard = wb["Dashboard"]

    category_ws = wb["By Category"]
    if category_ws.max_row >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Sales by Category"
        chart.y_axis.title = "Sales"
        chart.x_axis.title = "Category"
        data = Reference(category_ws, min_col=2, min_row=1, max_row=category_ws.max_row)
        categories = Reference(category_ws, min_col=1, min_row=2, max_row=category_ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7.5
        chart.width = 12
        dashboard.add_chart(chart, "F4")

    market_ws = wb["By Market"]
    if market_ws.max_row >= 2:
        max_row = min(market_ws.max_row, 11)
        chart = BarChart()
        chart.type = "bar"
        chart.style = 11
        chart.title = "Top Markets by Sales"
        chart.x_axis.title = "Sales"
        chart.y_axis.title = "Market"
        data = Reference(market_ws, min_col=2, min_row=1, max_row=max_row)
        categories = Reference(market_ws, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 12
        dashboard.add_chart(chart, "F20")


def build_report(
    current_file: Path,
    current_df: pd.DataFrame,
    rejected: pd.DataFrame,
    quality: pd.DataFrame,
    previous_file: Path | None,
    previous_df: pd.DataFrame | None,
    output_path: Path,
) -> None:
    current_metrics = calculate_metrics(current_df)
    previous_metrics = calculate_metrics(previous_df) if previous_df is not None and not previous_df.empty else None

    category = grouped_summary(current_df, "category")
    market = grouped_summary(current_df, "market")
    region = grouped_summary(current_df, "region")
    products = top_products(current_df)

    wb = Workbook()
    create_dashboard(
        wb,
        current_file,
        current_df,
        current_metrics,
        previous_file,
        previous_metrics,
    )

    write_dataframe_sheet(wb, "By Category", category, "CategoryTable")
    write_dataframe_sheet(wb, "By Market", market, "MarketTable")
    write_dataframe_sheet(wb, "By Region", region, "RegionTable")
    write_dataframe_sheet(wb, "Top Products", products, "TopProductsTable")
    write_dataframe_sheet(wb, "Data Quality", quality, "QualityTable")
    write_dataframe_sheet(wb, "Clean Data", current_df, "CleanDataTable")

    if not rejected.empty:
        write_dataframe_sheet(wb, "Rejected Rows", rejected, "RejectedRowsTable")

    add_dashboard_charts(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    configure_logging(DEFAULT_LOG_DIR)

    try:
        files = find_weekly_files(args.input_dir)
        current_file = files[-1]
        previous_file = files[-2] if len(files) >= 2 else None

        logging.info("Current weekly file: %s", current_file)
        current_raw = read_csv_flexible(current_file)
        current_df, rejected, quality = clean_sales_data(current_raw, current_file.name)

        previous_df: pd.DataFrame | None = None
        if previous_file is not None:
            logging.info("Previous weekly file: %s", previous_file)
            previous_raw = read_csv_flexible(previous_file)
            previous_df, _, _ = clean_sales_data(previous_raw, previous_file.name)

        report_date = current_df["order_date"].max().strftime("%Y-%m-%d")
        output_path = args.output_dir / f"weekly_sales_report_{report_date}.xlsx"

        build_report(
            current_file=current_file,
            current_df=current_df,
            rejected=rejected,
            quality=quality,
            previous_file=previous_file,
            previous_df=previous_df,
            output_path=output_path,
        )

        logging.info("Report created: %s", output_path)
        print(f"Report created: {output_path}")
    except Exception as error:
        logging.exception("Report generation failed: %s", error)
        raise SystemExit(1) from error


if __name__ == "__main__":
    main()
